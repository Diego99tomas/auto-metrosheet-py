from model_db.create_db import Create_DB
from openpyxl import load_workbook


EXCEL_NAME="interpolation/interpolation.xlsx"
DB_NAME="interpolation.db"


class ExportDataInterpFromExcel():
    
    TABLAS_MAP = {
    "VOLTAJE DC": "voltaje_dc",
    "CORRIENTE DC": "corriente_dc",
    "RESISTENCIAS": "resistencias",
    "VOLTAJE AC": "voltaje_ac",
    "CORRIENTE AC": "corriente_ac"}

    
    def migrar_datos(self,conn):
        """Migrar datos de interpolacion en excel a Base de Datos"""
        try:
            wb = load_workbook(filename=EXCEL_NAME, data_only=True)
        except FileNotFoundError:
            print("Error: No se encontró el Excel.")
            return

        cursor = conn.cursor()

        for hoja_excel, tabla_sql in self.TABLAS_MAP.items():
            if hoja_excel not in wb.sheetnames:
                print(f"Saltando {hoja_excel}: No existe en el Excel.")
                continue
                
            ws = wb[hoja_excel]
            es_ac = "ac" in tabla_sql
            
          
            cursor.execute(f"DELETE FROM {tabla_sql}")

          
            for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):

                if row[0] is None: continue  # Saltar filas vacías
                valor,unidad=row[0], row[1]
                lectura=Create_DB.normalizar_unidad(valor,unidad)

                if es_ac:
                    # row[0]=lectura, row[1]=unidad, row[2]=frecuencia, row[3]=medida, row[4]=incert
                    frec=row[2]
                    medida=Create_DB.normalizar_unidad(row[3],row[1])
                    incert=Create_DB.normalizar_unidad(row[4],row[1])
                    
                else:
                    # row[0]=lectura, row[1]=unidad, row[2]=medida, row[3]=incert
                    frec=None
                    medida=Create_DB.normalizar_unidad(row[2],row[1])
                    incert=Create_DB.normalizar_unidad(row[3],row[1])
                    
                sql=f"INSERT INTO {tabla_sql} (lectura, unidad, frecuencia, medida, incert) VALUES (?, ?, ?, ?, ?)" 
                params= (lectura,unidad,frec,medida,incert)
                 

                cursor.execute(sql, params)
            
        conn.commit() 
        print("Migración completada con éxito.")
        

if __name__ == "__main__":
    interp_db=Create_DB(DB_NAME=DB_NAME,EXCEL_NAME=EXCEL_NAME)
    conn=interp_db.create_db('lectura','unidad','frecuencia','medida','incert')
    export_data=ExportDataInterpFromExcel()
    export_data.migrar_datos(conn)
    conn.close()





