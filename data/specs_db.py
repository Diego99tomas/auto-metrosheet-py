from openpyxl import load_workbook
from model_db.create_db import Create_DB
from model_db.constants import Functions

EXCEL_NAME="specs/specs.xlsx"
DB_NAME="specs/specs.db"

class ExportDataSpecsFromExcel():
    

    def migrar_datos(self,conn):
        """Migrar datos de show_specs en excel a Base de Datos"""
        try:
            wb = load_workbook(filename=EXCEL_NAME, data_only=True)
        except FileNotFoundError:
            print("Error: No se encontró el Excel.")
            return

        cursor = conn.cursor()

        for func in Functions:
            if func.name not in wb.sheetnames:
                print(f"Saltando {func.name}: No existe en el Excel.")
                continue
                
            ws = wb[func.name]
            es_ac = "ac" in func.value
            
            # Limpiar tabla antes de insertar para evitar duplicados si re-corres el script
            cursor.execute(f"DELETE FROM {func.value}")

           
            for row in ws.iter_rows(min_row=1, max_row=70, values_only=True):
                if row[0] is None: continue  # Saltar filas vacías

                punto,unidad = row[0],row[1]
                valor=Create_DB.normalizar_unidad(punto,unidad)
                

                if es_ac:
                    frec=row[2]
                    decimal=Create_DB.normalizar_unidad(row[3],unidad)
                    show_spec=row[4]
                    lcomp=row[5] if len(row)>5 else None
                    
                else:
                    frec=None
                    lcomp=None
                    decimal=Create_DB.normalizar_unidad(row[2],unidad)
                    show_spec=row[3]
                    
                    
                sql=f"INSERT INTO {func.value} (valor, unidad, frecuencia, decimales, show_spec,lcomp) VALUES (?, ?, ?, ?, ?, ?)"   
                params=(valor,unidad,frec,decimal,show_spec,lcomp) 

                cursor.execute(sql,params)

        conn.commit()
        print("Migración completada exitosamente.")

if __name__ == "__main__":
    db_specs=Create_DB(DB_NAME=DB_NAME,EXCEL_NAME=EXCEL_NAME)
    conn=db_specs.create_db('valor','unidad','frecuencia','decimales','show_spec','lcomp')
    export_data=ExportDataSpecsFromExcel()
    export_data.migrar_datos(conn)
    conn.close()
     