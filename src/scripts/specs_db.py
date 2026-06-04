from database.models import Create_DB
from config.constants import Functions
from utility.data_processing import normalizar_unidad
from openpyxl import load_workbook

EXCEL_NAME=r"data\raw\specs.xlsx"
DB_NAME=r"data\sqlite\specs.db"


def migrar_datos(conn):
    """Migrar datos de show_specs en excel a Base de Datos"""

    try:
        wb = load_workbook(filename=EXCEL_NAME, data_only=True)
    except FileNotFoundError:
        print("Error: No se encontró el Excel.")
        return

    cursor = conn.cursor()

    for func in Functions:
        if func.name not in wb.sheetnames:
            print(f"Saltando {func.name}: No existe en el Excel.")
            continue
            
        ws = wb[func.name]
        es_ac = "ac" in func.value
        
        cursor.execute(f"DELETE FROM {func.value}")

        
        for row in ws.iter_rows(min_row=1, max_row=70, values_only=True):
            if row[0] is None: continue  # Saltar filas vacías

            punto,unidad = row[0],row[1]
            valor=normalizar_unidad(punto,unidad)
            

            if es_ac:
                frec=row[2]
                decimal=normalizar_unidad(row[3],unidad)
                show_spec=row[4]
                lcomp=row[5] if len(row)>5 else None
                
            else:
                frec=None
                lcomp=None
                decimal=normalizar_unidad(row[2],unidad)
                show_spec=row[3]
                
                
            sql=f"INSERT INTO {func.value} (valor, unidad, frecuencia, decimales, show_spec,lcomp) VALUES (?, ?, ?, ?, ?, ?)"   
            params=(valor,unidad,frec,decimal,show_spec,lcomp) 

            cursor.execute(sql,params)

    conn.commit()
    print("Migración completada exitosamente.")

if __name__ == "__main__":
    db_specs=Create_DB(DB_NAME,EXCEL_NAME)
    conn=db_specs.create_db('valor','unidad','frecuencia','decimales','show_spec','lcomp')
    migrar_datos(conn)
    conn.close()
     