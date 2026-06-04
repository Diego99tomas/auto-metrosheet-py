import os
import sqlite3
from openpyxl import load_workbook

from src.database.models import DataBaseBuilder
from src.database.connection import INTERPOLATION_DB_PATH
from src.database.connection import get_connection
from src.config.constants import Functions
from src.utility.data_processing import normalizar_unidad


EXCEL_NAME=os.path.join("data","raw","interpolation.xlsx")

def migrar_datos():
    """Migrar datos de interpolacion en excel a Base de Datos"""

    try:
        wb = load_workbook(filename=EXCEL_NAME, data_only=True)
    except FileNotFoundError:
        print("Error: No se encontró el Excel.")
        return
    
    conn=get_connection(INTERPOLATION_DB_PATH)
    cursor = conn.cursor()

    try:
        for func in Functions:
            if func.name not in wb.sheetnames:
                print(f"Saltando {func.name}: No existe en el Excel.")
                continue
                
            ws = wb[func.name]
            es_ac = "ac" in func.value
                    
            cursor.execute(f"DELETE FROM {func.value}")

            for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):

                if row[0] is None: continue  # Saltar filas vacías
                valor,unidad=row[0], row[1]
                lectura=normalizar_unidad(valor,unidad)

                if es_ac:
                    # row[0]=lectura, row[1]=unidad, row[2]=frecuencia, row[3]=medida, row[4]=incert
                    frec=row[2]
                    medida=normalizar_unidad(row[3],row[1])
                    incert=normalizar_unidad(row[4],row[1])
                    
                else:
                    # row[0]=lectura, row[1]=unidad, row[2]=medida, row[3]=incert
                    frec=None
                    medida=normalizar_unidad(row[2],row[1])
                    incert=normalizar_unidad(row[3],row[1])
                    
                sql=f"INSERT INTO {func.value} (lectura, unidad, frecuencia, medida, incert) VALUES (?, ?, ?, ?, ?)" 
                params= (lectura,unidad,frec,medida,incert)
                    
                cursor.execute(sql, params)
            
        conn.commit() 
        print("Migración completada con éxito.")
    
    except sqlite3.Error as e:
        conn.rollback()
        print(f"Error de SQlite durante migracion: {e}")    
    
    except Exception as e:
        conn.rollback()
        print(f"Error inesperado al procesar datos: {e}")
    
    finally:
        conn.close()   

if __name__ == "__main__":
    builder=DataBaseBuilder(INTERPOLATION_DB_PATH)
    conn=builder.create_schema('lectura','unidad','frecuencia','medida','incert')
    migrar_datos()
    





