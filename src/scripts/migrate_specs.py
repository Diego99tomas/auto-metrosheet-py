import os
import sqlite3
from openpyxl import load_workbook

from src.database.models import DataBaseBuilder
from src.database.connection import get_connection,SPECS_DB_PATH
from config.constants import Functions
from utility.data_processing import normalizar_unidad
from utility.excel_tools import load_template_excel


EXCEL_NAME=os.path.join("data","raw","specs.xlsx")


def migrar_datos():
    """Migrar datos de show_specs en excel a Base de Datos"""

    try:
        wb = load_workbook(filename=EXCEL_NAME, data_only=True)
    except FileNotFoundError:
        print("Error: No se encontró el Excel.")
        return
    
    conn=get_connection(SPECS_DB_PATH)
    cursor = conn.cursor()

    try:

        for func in Functions:
            if func.name not in wb.sheetnames:
                print(f"Saltando {func.name}: No existe en el Excel.")
                continue
                
            ws = wb[func.name]
            es_ac = "ac" in func.value
            
            cursor.execute(f"DELETE FROM {func.value}")

            
            for row in ws.iter_rows(min_row=1, max_row=70, values_only=True):
                if row[0] is None: continue  # Saltar filas vacías

                punto,unidad = row[0],row[1]
                valor=normalizar_unidad(punto,unidad)
                

                if es_ac:
                    frec=row[2]
                    decimal=normalizar_unidad(row[3],unidad)
                    show_spec=row[4]
                    lcomp=row[5] if len(row)>5 else None
                    
                else:
                    frec=None
                    lcomp=None
                    decimal=normalizar_unidad(row[2],unidad)
                    show_spec=row[3]
                    
                    
                sql=f"INSERT INTO {func.value} (valor, unidad, frecuencia, decimales, show_spec,lcomp) VALUES (?, ?, ?, ?, ?, ?)"   
                params=(valor,unidad,frec,decimal,show_spec,lcomp) 

                cursor.execute(sql,params)

        conn.commit()
        print("Migración completada exitosamente.")

    except sqlite3.Error as e:
        conn.rollback()
        print(f"Error de SQlite durante migracion: {e}")    
    
    except Exception as e:
        conn.rollback()
        print(f"Error inesperado al procesar datos: {e}")
    
    finally:
        conn.close() 

if __name__ == "__main__":
    db_specs=DataBaseBuilder(SPECS_DB_PATH)
    db_specs.create_schema('valor','unidad','frecuencia','decimales','show_spec','lcomp')
    migrar_datos()
    
     